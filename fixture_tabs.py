import tkinter as tk
import openpyxl
from tkinter import ttk, messagebox, filedialog
from openpyxl.styles import Alignment
from db_helper import get_conn
from fixture_helper import generate_part_no_by_category
from fixture_helper import (
    insert_fixture,
    delete_fixture,
    add_stock,
    transfer_stock,
    update_fixture,
    adjust_stock,
    get_overview_by_warehouse,
    generate_location,
    get_fixture_by_part_no,
    validate_location,
    CORE_WAREHOUSES
)

EXCHANGE_RATE = 30.375
CATEGORIES = ["電腦設備類", "載具類", "治具類", "板子類", "主機類", "電供類", "線材類", "卡片類", "其它類", "消耗類"]

WAREHOUSES = CORE_WAREHOUSES

def build_fixture_tab(parent, current_user: str = None, on_change=None):
    frames, trees, total_labels, total_price_labels = {}, {}, {}, {}

    def _normalize_username(u):
        if u is None:
            return ""
        if isinstance(u, str):
            return u.strip()
        if isinstance(u, dict):
            v = u.get("user") or u.get("username") or ""
            return str(v).strip()
        return str(u).strip()

    username = _normalize_username(current_user)

    def _get_user_permissions(user_name: str):
        flags = {
            "can_view_fixture": 0,
            "can_edit_fixture": 0,
            "can_adjust_fixture": 0,
            "can_view_fixture_logs": 0,
            "can_delete_fixture_logs": 0
        }
        if not user_name:
            return flags
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT "
                "COALESCE(can_view_fixture, 0), "
                "COALESCE(can_edit_fixture, 0), "
                "COALESCE(can_adjust_fixture, 0), "
                "COALESCE(can_view_fixture_logs, 0), "
                "COALESCE(can_delete_fixture_logs, 0) "
                "FROM users WHERE username=?",
                (user_name,),
            )
            row = cur.fetchone()
            if not row:
                return flags
            flags["can_view_fixture"] = int(row[0] or 0)
            flags["can_edit_fixture"] = int(row[1] or 0)
            flags["can_adjust_fixture"] = int(row[2] or 0)
            flags["can_view_fixture_logs"] = int(row[3] or 0)
            flags["can_delete_fixture_logs"] = int(row[4] or 0)
            return flags

    perms = _get_user_permissions(username)

    if int(perms.get("can_view_fixture") or 0) != 1:
        holder = ttk.Frame(parent)
        holder.pack(fill="both", expand=True)
        ttk.Label(holder, text="無權限使用治具管理").pack(padx=10, pady=10)
        return

    notebook = ttk.Notebook(parent)
    notebook.pack(fill="both", expand=True)

    form = ttk.LabelFrame(parent, text="治具操作")
    form.pack(fill="x", padx=5, pady=5)

    filter_frame = ttk.LabelFrame(parent, text="查詢 / 篩選")
    filter_frame.pack(fill="x", padx=5, pady=5)

    filter_part_no_var = tk.StringVar()
    filter_name_var = tk.StringVar()
    filter_spec_var = tk.StringVar()
    filter_group_var = tk.StringVar()
    apply_all_wh_var = tk.IntVar(value=0)

    ttk.Label(filter_frame, text="料號關鍵字:").grid(row=0, column=0, sticky="e", padx=3, pady=2)
    entry_filter_part_no = ttk.Entry(filter_frame, textvariable=filter_part_no_var, width=18)
    entry_filter_part_no.grid(row=0, column=1, sticky="w", padx=3, pady=2)

    ttk.Label(filter_frame, text="品名關鍵字:").grid(row=0, column=2, sticky="e", padx=3, pady=2)
    entry_filter_name = ttk.Entry(filter_frame, textvariable=filter_name_var, width=18)
    entry_filter_name.grid(row=0, column=3, sticky="w", padx=3, pady=2)

    ttk.Label(filter_frame, text="規格關鍵字:").grid(row=0, column=4, sticky="e", padx=3, pady=2)
    entry_filter_spec = ttk.Entry(filter_frame, textvariable=filter_spec_var, width=18)
    entry_filter_spec.grid(row=0, column=5, sticky="w", padx=3, pady=2)

    ttk.Label(filter_frame, text="類群:").grid(row=0, column=6, sticky="e", padx=3, pady=2)
    combo_filter_group = ttk.Combobox(filter_frame, textvariable=filter_group_var, values=[""] + CATEGORIES, state="readonly", width=16)
    combo_filter_group.grid(row=0, column=7, sticky="w", padx=3, pady=2)

    chk_apply_all = ttk.Checkbutton(filter_frame, text="套用全部倉別", variable=apply_all_wh_var)
    chk_apply_all.grid(row=0, column=8, sticky="w", padx=8, pady=2)

    labels = ["治具料號", "治具品名", "治具規格", "治具類群", "治具單價(USD)", "安全庫存", "儲位", "入庫數量"]
    entries = {}
    for i, lab in enumerate(labels):
        ttk.Label(form, text=lab + ":").grid(row=i, column=0, sticky="e", padx=3, pady=2)
        ent = ttk.Entry(form, width=20)
        ent.grid(row=i, column=1, sticky="w", padx=3, pady=2)
        entries[lab] = ent

    combo_cat = ttk.Combobox(form, values=CATEGORIES, state="readonly", width=18)
    combo_cat.grid(row=3, column=1, sticky="w", padx=3, pady=2)
    entries["治具類群"] = combo_cat

    def _current_wh():
        try:
            return notebook.tab(notebook.select(), "text")
        except Exception:
            return ""

    def _get_filters():
        return {
            "part_no": (filter_part_no_var.get() or "").strip(),
            "name": (filter_name_var.get() or "").strip(),
            "spec": (filter_spec_var.get() or "").strip(),
            "group": (filter_group_var.get() or "").strip(),
        }

    def refresh_current():
        wh = _current_wh()
        if not wh:
            return
        f = _get_filters()
        refresh_fixture_tree(trees[wh], wh, total_labels[wh], total_price_labels[wh], filters=f)

    def refresh_all():
        f = _get_filters()
        for wh in WAREHOUSES:
            refresh_fixture_tree(trees[wh], wh, total_labels[wh], total_price_labels[wh], filters=f)

    def on_apply_filter():
        if int(apply_all_wh_var.get() or 0) == 1:
            refresh_all()
        else:
            refresh_current()

    def on_reset_filter():
        filter_part_no_var.set("")
        filter_name_var.set("")
        filter_spec_var.set("")
        filter_group_var.set("")
        apply_all_wh_var.set(0)
        refresh_all()

    btn_filter_apply = ttk.Button(filter_frame, text="查詢", command=on_apply_filter)
    btn_filter_reset = ttk.Button(filter_frame, text="重置", command=on_reset_filter)
    btn_filter_apply.grid(row=0, column=9, padx=6, pady=2)
    btn_filter_reset.grid(row=0, column=10, padx=6, pady=2)

    def _bind_enter_to_apply(widget):
        try:
            widget.bind("<Return>", lambda e: on_apply_filter())
        except Exception:
            pass

    _bind_enter_to_apply(entry_filter_part_no)
    _bind_enter_to_apply(entry_filter_name)
    _bind_enter_to_apply(entry_filter_spec)

    def after_action():
        if int(apply_all_wh_var.get() or 0) == 1:
            refresh_all()
        else:
            refresh_current()
        if on_change:
            on_change()

    def _require_edit():
        if int(perms.get("can_edit_fixture") or 0) != 1:
            messagebox.showerror("錯誤", "無治具編輯權限")
            return False
        return True

    def _require_adjust():
        if int(perms.get("can_adjust_fixture") or 0) != 1:
            messagebox.showerror("錯誤", "無治具調帳權限")
            return False
        return True

    def on_generate_part_no():
        if not _require_edit():
            return
        cat = combo_cat.get().strip()
        if not cat:
            messagebox.showerror("錯誤", "請先選擇治具類群")
            return
        try:
            part_no = generate_part_no_by_category(cat)
            entries["治具料號"].delete(0, tk.END)
            entries["治具料號"].insert(0, part_no)
            messagebox.showinfo("完成", f"自動生成料號：{part_no}")
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    def on_add_fixture():
        if not _require_edit():
            return
        part_no = entries["治具料號"].get().strip()
        part_name = entries["治具品名"].get().strip()
        part_spec = entries["治具規格"].get().strip()
        part_group = combo_cat.get().strip()
        usd = entries["治具單價(USD)"].get().strip()
        safety = entries["安全庫存"].get().strip()
        loc_raw = entries["儲位"].get().strip()

        if not (part_no.isdigit() and len(part_no) in (8, 12)):
            messagebox.showerror("錯誤", "治具料號必須為 8 或 12 碼數字")
            return
        if not part_name or not part_spec or not part_group:
            messagebox.showerror("錯誤", "治具品名/規格/類群不可為空")
            return
        try:
            usd_val = float(usd or 0)
            usd_val = int(usd_val * 1000) / 1000.0
            ntd_val = int(usd_val * EXCHANGE_RATE * 1000) / 1000.0
            safety_val = int(safety or 0)
        except:
            messagebox.showerror("錯誤", "單價或安全庫存格式錯誤")
            return

        try:
            loc_fmt = validate_location(part_no, loc_raw)
        except Exception as e:
            messagebox.showerror("錯誤", str(e))
            return

        try:
            insert_fixture(
                part_no, part_name, part_spec, part_group,
                ntd_val, safety_val, loc_fmt,
                unit_price_usd=usd_val,
                user=username
            )
            messagebox.showinfo("完成", f"治具 {part_no} 已新增")
            after_action()
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    def on_delete_fixture():
        if not _require_edit():
            return
        part_no = entries["治具料號"].get().strip()
        if not part_no:
            return
        if not messagebox.askyesno("確認", f"確定刪除 {part_no}？將移除所有倉別存量與 BOM 及相關紀錄"):
            return
        try:
            delete_fixture(part_no, user=username)
            messagebox.showinfo("完成", f"{part_no} 已刪除")
            after_action()
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    def on_add_stock():
        if not _require_edit():
            return
        part_no = entries["治具料號"].get().strip()
        qty_str = entries["入庫數量"].get().strip()
        try:
            usable_qty = int(qty_str)
        except:
            messagebox.showerror("錯誤", "入庫量必須是整數")
            return
        if not (part_no.isdigit() and len(part_no) in (8, 12)):
            messagebox.showerror("錯誤", "治具料號必須為 8 或 12 碼數字")
            return
        current_tab = notebook.tab(notebook.select(), "text")
        wh = current_tab
        try:
            add_stock(part_no, usable_qty, wh, user=username)
            messagebox.showinfo("完成", f"{part_no} 已入庫 {usable_qty} 至 {wh}")
            after_action()
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    def on_transfer():
        if not _require_edit():
            return
        part_no = entries["治具料號"].get().strip()
        qty_str = entries["入庫數量"].get().strip()
        from_wh, to_wh = combo_from.get(), combo_to.get()
        try:
            usable_qty = int(qty_str)
        except:
            messagebox.showerror("錯誤", "調撥量必須是整數")
            return
        if not (part_no.isdigit() and len(part_no) in (8, 12)):
            messagebox.showerror("錯誤", "治具料號必須為 8 或 12 碼數字")
            return
        if from_wh == to_wh:
            messagebox.showerror("錯誤", "來源與目標倉別相同")
            return
        try:
            transfer_stock(part_no, usable_qty, from_wh, to_wh, user=username)
            messagebox.showinfo("完成", f"{part_no} 已調撥 {usable_qty} 從 {from_wh} 到 {to_wh}")
            after_action()
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    def on_update_fixture():
        if not _require_edit():
            return
        part_no = entries["治具料號"].get().strip()
        part_name = entries["治具品名"].get().strip()
        part_spec = entries["治具規格"].get().strip()
        part_group = combo_cat.get().strip()
        usd = entries["治具單價(USD)"].get().strip()
        loc_raw = entries["儲位"].get().strip()
        safety = entries["安全庫存"].get().strip()
        try:
            usd_val = float(usd or 0)
            usd_val = int(usd_val * 1000) / 1000.0
            ntd_val = int(usd_val * EXCHANGE_RATE * 1000) / 1000.0
            safety_val = int(safety or 0)
        except:
            messagebox.showerror("錯誤", "單價或安全庫存格式錯誤")
            return
        try:
            loc_fmt = validate_location(part_no, loc_raw)
        except Exception as e:
            messagebox.showerror("錯誤", str(e))
            return
        try:
            update_fixture(
                part_no,
                part_name=part_name,
                part_spec=part_spec,
                part_group=part_group,
                unit_price_ntd=ntd_val,
                unit_price_usd=usd_val,
                safety_stock=safety_val,
                storage_location=loc_fmt,
                user=username
            )
            messagebox.showinfo("完成", f"{part_no} 的資料已更新")
            after_action()
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    def on_export_excel():
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file:
            return
        wb = openpyxl.Workbook()
        for wh in WAREHOUSES:
            ws = wb.create_sheet(title=wh)
            headers = [
                "治具料號", "治具品名", "治具規格", "治具類群",
                "單價USD", "單價NTD", "總價NTD",
                "儲位", "安全庫存", "可用數量",
                "預估請購量", "預估請購金額"
            ]
            ws.append(headers)
            rows = get_overview_by_warehouse(wh)

            total_estimate_cost = 0
            for (part_no, part_name, part_spec, part_group, unit_price_usd, unit_price_ntd, safety_stock, storage_location, usable_qty) in rows:
                unit_price_usd = int((unit_price_usd or 0) * 1000) / 1000.0
                total_price_item = int((unit_price_ntd or 0) * (usable_qty or 0) * 1000) / 1000.0

                if (safety_stock or 0) > (usable_qty or 0):
                    est_qty = (safety_stock or 0) - (usable_qty or 0)
                else:
                    est_qty = 0
                est_cost = int(est_qty * (unit_price_ntd or 0) * 1000) / 1000.0
                total_estimate_cost += est_cost

                ws.append([
                    part_no, part_name, part_spec, part_group,
                    unit_price_usd, unit_price_ntd, total_price_item,
                    storage_location if wh == "虹堡" else "", safety_stock if wh == "虹堡" else 0, usable_qty,
                    est_qty, est_cost
                ])

            ws.append([])
            ws.append([""] * 10 + ["預估請購總金額", total_estimate_cost])

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            col_widths = {
                "A": 20, "B": 50, "C": 50, "D": 15,
                "E": 12, "F": 12, "G": 18,
                "H": 15, "I": 12, "J": 12,
                "K": 20, "L": 18,
            }
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            last_row = ws.max_row
            ws.cell(row=last_row, column=11).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=last_row, column=12).alignment = Alignment(horizontal="center", vertical="center")

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(file)
        messagebox.showinfo("完成", f"已匯出 Excel：{file}")

    def on_generate_location():
        if not _require_edit():
            return
        prefix = entries["儲位"].get().strip()
        try:
            loc = generate_location(prefix if prefix else "")
            entries["儲位"].delete(0, tk.END)
            entries["儲位"].insert(0, loc)
            messagebox.showinfo("完成", f"自動生成儲位：{loc}")
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    btn_add = ttk.Button(form, text="建立治具", command=on_add_fixture)
    btn_del = ttk.Button(form, text="刪除治具", command=on_delete_fixture)
    btn_upd = ttk.Button(form, text="修改資料", command=on_update_fixture)
    btn_gen_no = ttk.Button(form, text="生成料號", command=on_generate_part_no)
    btn_gen_loc = ttk.Button(form, text="生成儲位", command=on_generate_location)
    btn_in = ttk.Button(form, text="入庫", command=on_add_stock)

    btn_add.grid(row=0, column=2, padx=5)
    btn_del.grid(row=1, column=2, padx=5)
    btn_upd.grid(row=6, column=2, padx=5)
    btn_gen_no.grid(row=0, column=3, padx=5)
    btn_gen_loc.grid(row=6, column=3, padx=5)
    btn_in.grid(row=7, column=2, padx=5)

    transfer_frame = ttk.LabelFrame(form, text="調撥 / 其它操作")
    transfer_frame.grid(row=8, column=0, columnspan=4, sticky="ew", pady=5)

    ttk.Label(transfer_frame, text="來源:").grid(row=0, column=0, padx=3)
    ttk.Label(transfer_frame, text="目標:").grid(row=0, column=2, padx=3)

    combo_from = ttk.Combobox(transfer_frame, values=CORE_WAREHOUSES, state="readonly", width=12)
    combo_to = ttk.Combobox(transfer_frame, values=CORE_WAREHOUSES, state="readonly", width=12)
    combo_from.grid(row=0, column=1, padx=3)
    combo_to.grid(row=0, column=3, padx=3)

    btn_transfer = ttk.Button(transfer_frame, text="執行調撥", command=on_transfer)
    btn_export = ttk.Button(transfer_frame, text="匯出Excel", command=on_export_excel)
    btn_transfer.grid(row=1, column=0, pady=3, padx=5)
    btn_export.grid(row=1, column=2, pady=3, padx=5)

    adjust_frame = ttk.LabelFrame(form, text="調帳")
    adjust_frame.grid(row=9, column=0, columnspan=4, sticky="ew", pady=5)

    ttk.Label(adjust_frame, text="倉別:").grid(row=0, column=0, padx=3, pady=2, sticky="e")
    adj_wh_var = tk.StringVar()
    combo_adj_wh = ttk.Combobox(adjust_frame, textvariable=adj_wh_var, values=CORE_WAREHOUSES, state="readonly", width=12)
    combo_adj_wh.grid(row=0, column=1, padx=3, pady=2, sticky="w")

    ttk.Label(adjust_frame, text="模式:").grid(row=0, column=2, padx=3, pady=2, sticky="e")
    adj_mode_var = tk.StringVar()
    combo_adj_mode = ttk.Combobox(adjust_frame, textvariable=adj_mode_var, values=("差額", "盤點"), state="readonly", width=12)
    combo_adj_mode.grid(row=0, column=3, padx=3, pady=2, sticky="w")
    combo_adj_mode.set("差額")

    ttk.Label(adjust_frame, text="數量:").grid(row=1, column=0, padx=3, pady=2, sticky="e")
    adj_qty_var = tk.StringVar()
    entry_adj_qty = ttk.Entry(adjust_frame, textvariable=adj_qty_var, width=14)
    entry_adj_qty.grid(row=1, column=1, padx=3, pady=2, sticky="w")

    ttk.Label(adjust_frame, text="原因:").grid(row=1, column=2, padx=3, pady=2, sticky="e")
    adj_reason_var = tk.StringVar()
    entry_adj_reason = ttk.Entry(adjust_frame, textvariable=adj_reason_var, width=24)
    entry_adj_reason.grid(row=1, column=3, padx=3, pady=2, sticky="w")

    def on_adjust():
        if not _require_adjust():
            return
        part_no = entries["治具料號"].get().strip()
        wh = adj_wh_var.get().strip()
        mode = adj_mode_var.get().strip()
        qty_text = adj_qty_var.get().strip()
        reason = adj_reason_var.get().strip()

        if not (part_no.isdigit() and len(part_no) in (8, 12)):
            messagebox.showerror("錯誤", "治具料號必須為 8 或 12 碼數字")
            return
        if not wh:
            messagebox.showerror("錯誤", "請選擇倉別")
            return
        if not qty_text:
            messagebox.showerror("錯誤", "請輸入調帳數量")
            return
        if reason.strip() == "":
            messagebox.showerror("錯誤", "請輸入調帳原因")
            return
        try:
            qty_val = int(qty_text)
        except:
            messagebox.showerror("錯誤", "調帳數量必須是整數")
            return

        try:
            adjust_stock(part_no, wh, mode, qty_val, reason, user=username)
            messagebox.showinfo("完成", f"{part_no} 已完成調帳（{wh}）")
            adj_qty_var.set("")
            adj_reason_var.set("")
            after_action()
        except Exception as e:
            messagebox.showerror("錯誤", str(e))

    btn_adjust = ttk.Button(adjust_frame, text="執行調帳", command=on_adjust)
    btn_adjust.grid(row=0, column=4, rowspan=2, padx=6, pady=2)

    can_edit = int(perms.get("can_edit_fixture") or 0) == 1
    can_adjust = int(perms.get("can_adjust_fixture") or 0) == 1

    if not can_edit:
        btn_add.config(state="disabled")
        btn_del.config(state="disabled")
        btn_upd.config(state="disabled")
        btn_gen_no.config(state="disabled")
        btn_gen_loc.config(state="disabled")
        btn_in.config(state="disabled")
        btn_transfer.config(state="disabled")
        combo_from.config(state="disabled")
        combo_to.config(state="disabled")

    if not can_adjust:
        btn_adjust.config(state="disabled")
        combo_adj_wh.config(state="disabled")
        combo_adj_mode.config(state="disabled")
        entry_adj_qty.config(state="disabled")
        entry_adj_reason.config(state="disabled")

    for wh in WAREHOUSES:
        frame = ttk.Frame(notebook)
        frames[wh] = frame
        notebook.add(frame, text=wh)
        top = ttk.Frame(frame)
        top.pack(fill="x")
        total_labels[wh] = ttk.Label(top, text="合計可用數量: 0")
        total_labels[wh].pack(side="right", padx=6)
        total_price_labels[wh] = ttk.Label(top, text="倉別總價: 0 NTD")
        total_price_labels[wh].pack(side="right", padx=6)

        cols = ("part_no", "part_name", "part_spec", "part_group", "unit_price_usd", "unit_price_ntd", "total_price_ntd", "safety_stock", "storage_location", "usable_qty")
        headers = ["治具料號", "治具品名", "治具規格", "治具類群", "單價USD", "單價NTD", "總價NTD", "安全庫存", "儲位", "可用數量"]

        tree = ttk.Treeview(frame, columns=cols, show="headings")
        trees[wh] = tree
        for c, t in zip(cols, headers):
            tree.heading(c, text=t)
            tree.column(c, width=90, anchor="center")
        tree.column("part_name", width=240)
        tree.column("part_spec", width=240)
        tree.pack(fill="both", expand=True)
        refresh_fixture_tree(tree, wh, total_labels[wh], total_price_labels[wh], filters=_get_filters())

        def on_double(event, tree=tree, wh=wh):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            entries["治具料號"].delete(0, tk.END)
            entries["治具料號"].insert(0, vals[0])
            entries["治具品名"].delete(0, tk.END)
            entries["治具品名"].insert(0, vals[1])
            entries["治具規格"].delete(0, tk.END)
            entries["治具規格"].insert(0, vals[2])
            combo_cat.set(vals[3])
            entries["治具單價(USD)"].delete(0, tk.END)
            entries["治具單價(USD)"].insert(0, vals[4])
            entries["安全庫存"].delete(0, tk.END)
            entries["安全庫存"].insert(0, vals[7])
            entries["儲位"].delete(0, tk.END)
            entries["儲位"].insert(0, vals[8])
            combo_from.set(wh)
            adj_wh_var.set(wh)

        tree.bind("<Double-1>", on_double)

    def _on_tab_changed(event=None):
        if int(apply_all_wh_var.get() or 0) == 1:
            refresh_all()
        else:
            refresh_current()

    notebook.bind("<<NotebookTabChanged>>", _on_tab_changed)


def refresh_fixture_tree(tree, warehouse, total_label=None, total_price_label=None, filters=None):
    tree.delete(*tree.get_children())
    rows = get_overview_by_warehouse(warehouse)

    filters = filters or {}
    f_part_no = (filters.get("part_no") or "").strip().lower()
    f_name = (filters.get("name") or "").strip().lower()
    f_spec = (filters.get("spec") or "").strip().lower()
    f_group = (filters.get("group") or "").strip()

    total_qty = 0
    total_price = 0.0
    for (part_no, part_name, part_spec, part_group, unit_price_usd, unit_price_ntd, safety_stock, storage_location, usable_qty) in rows:
        part_no_s = "" if part_no is None else str(part_no)
        part_name_s = "" if part_name is None else str(part_name)
        part_spec_s = "" if part_spec is None else str(part_spec)
        part_group_s = "" if part_group is None else str(part_group)

        if f_part_no and f_part_no not in part_no_s.lower():
            continue
        if f_name and f_name not in part_name_s.lower():
            continue
        if f_spec and f_spec not in part_spec_s.lower():
            continue
        if f_group and f_group != part_group_s:
            continue

        unit_price_ntd = unit_price_ntd or 0
        usable_qty = usable_qty or 0
        total_price_item = int(unit_price_ntd * usable_qty * 1000) / 1000.0
        show_location = storage_location if warehouse == "虹堡" else ""
        show_safety = safety_stock if warehouse == "虹堡" else 0
        tree.insert(
            "",
            "end",
            values=(
                part_no,
                part_name,
                part_spec,
                part_group,
                unit_price_usd,
                unit_price_ntd,
                total_price_item,
                show_safety,
                show_location,
                usable_qty,
            ),
        )
        total_qty += usable_qty
        total_price += total_price_item

    if total_label is not None:
        total_label.config(text=f"合計可用數量: {total_qty}")
    if total_price_label is not None:
        total_price_label.config(text=f"倉別總價: {int(total_price * 1000) / 1000.0} NTD")